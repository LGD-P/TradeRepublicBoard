#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""TradeRepublicBoard — build a polished portfolio workbook from a Trade
Republic CSV export.

The whole workbook (styles, formulas, charts, sheets) is rebuilt from scratch
on every run, so no starting file is required. The script is self-contained and
can be shared as-is.

    python tr_board.py --fi "transactions.csv" --fo "board.xlsx" --en

Generated sheets
----------------
  * Read me      : short user guide
  * Dashboard    : KPIs, performance, month-by-month portfolio value + charts
  * Investments  : ETF journal (Buy / Saveback / Sell) + live prices
  * Yearly       : contributions and gains per year
  * By ETF       : performance per line + interactive per-year detail
  * Tax          : cash interest + realised capital gains (tax filing)
  * Stock Picking-IPO : single stocks (buy/sell, weighted-average cost)

Business rules
--------------
  * ETF portfolio  = transactions with asset_class == FUND.
  * Single stocks  = other TRADING transactions (routed to the stock sheet).
  * Saveback       = a BENEFITS_SAVEBACK credit reinvested by a buy of the same
                     amount; that buy is tagged "Saveback".
  * Card payments, transfers and marketing are ignored (cash interest is kept
    on the Tax sheet as it is taxable).
  * Realised gains use the weighted-average cost method.
  * Prices already typed into an existing --fo are preserved.

This is an unofficial personal tool and is not affiliated with Trade Republic.
"""

import argparse
import csv
import json
import os
import shutil
import urllib.request
from collections import Counter, defaultdict
from datetime import datetime

import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ==========================================================================
#  Internationalisation (en / fr)
# ==========================================================================
#
# Every user-facing string lives here. `S` holds the strings for the active
# language and `MONTHS` / `MONTHS_ABBR` the localised month names. A few strings
# (sheet names, the "Buy/Saveback/Sell" tags, the control words) are used BOTH
# as displayed text AND inside formulas, so they must stay consistent -- hence
# the single source of truth below.

MONTH_NAMES = {
    "en": ["January", "February", "March", "April", "May", "June", "July",
           "August", "September", "October", "November", "December"],
    "fr": ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet",
           "Août", "Septembre", "Octobre", "Novembre", "Décembre"],
}
MONTH_ABBR = {
    "en": ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep",
           "Oct", "Nov", "Dec"],
    "fr": ["Janv", "Févr", "Mars", "Avri", "Mai", "Juin", "Juil", "Août",
           "Sept", "Octo", "Nove", "Déce"],
}

TXT = {
    "en": {
        # Sheet names (also referenced in formulas)
        "sheet_readme": "Read me",
        "sheet_dashboard": "Dashboard",
        "sheet_investments": "Investments",
        "sheet_yearly": "Yearly",
        "sheet_by_etf": "By ETF",
        "sheet_tax": "Tax",
        "sheet_stock": "Stock Picking-IPO",
        # Transaction tags (also SUMIFS criteria)
        "buy": "Buy",
        "saveback": "Saveback",
        "sell": "Sell",
        "saveback_card": "Card saveback",
        "roundup_card": "Card round-up",
        # Control column
        "ok": "OK",
        "gap": "gap",
        "total": "TOTAL",
        "totals": "TOTALS",
        # Common column headers
        "col_date": "Date", "col_year": "Year", "col_month": "Month",
        "col_type": "Type", "col_name": "Name / ETF", "col_unit_price": "Unit price (€)",
        "col_shares": "Shares", "col_amount": "Total (€)", "col_control": "Check",
        "col_current_value": "Current value (€)", "col_fees": "Fees (€)",
        "col_comment": "Comment", "col_isin": "ISIN", "col_current_price": "Current price (€)",
        "col_updated": "Updated", "col_etf_stock": "ETF / stock",
        # Read me
        "rm_title": "Investment tracker — Trade Republic",
        "rm_blocks": [
            ("How it works", [
                "Each share you own = one journal row (Buy or Saveback), with its "
                "price, its shares and its total.",
                "Portfolio value = shares × current price. Cost basis = sum of totals.",
                "YELLOW cells are yours to fill (salmon rows = savebacks). "
                "Everything else is automatic."]),
            ("About fees", [
                "A row total = shares × price (what TR actually invests). One-off "
                "order fees go in the Fees column, separately."]),
            ("Performance and colours", [
                "\"Performance\" (Dashboard) = your market gain on all invested "
                "money, like TR shows it.",
                "\"Saveback contribution\" = today's value of the shares offered "
                "by savebacks.",
                "All gains/losses are green (gain) or red (loss)."]),
            ("The sheets", [
                "\"Investments\": live prices (ISIN auto-filled) + ETF journal. "
                "Consistency gaps turn orange.",
                "\"Yearly\": contributions and gains per year.",
                "\"By ETF\": performance per line, PLUS an interactive per-year "
                "detail (dropdown) with charts.",
                "\"Dashboard\": KPIs, gains, and a monthly detail tracing the "
                "portfolio value over time (each month priced at its own "
                "transactions, last point = current price).",
                "\"Tax\": cash interest and realised gains (filing).",
                "\"Stock Picking / IPO\": single stocks (buy/sell, "
                "weighted-average cost)."]),
            ("Monthly refresh", [
                "Re-export the CSV from Trade Republic and re-run the script: the "
                "journal is rebuilt and your prices are preserved.",
                "\"Auto prices\": run with --auto-prices to fetch prices by ISIN "
                "(Deutsche Börse / Xetra, Yahoo fallback); otherwise just fill the "
                "yellow Current price column.",
                "Personal tracking tool, not investment advice."]),
        ],
        # Investments
        "inv_title": "Investments",
        "inv_prices_note": "Current prices — one per ETF (source: your TR app). "
                           "ISIN auto-filled, you only type the price.",
        "inv_journal_note": "Journal — yellow to fill (salmon = saveback). "
                            "Total = shares × price; fees go in their own column. "
                            "Orange row = consistency gap to fix.",
        "inv_totals_buy_sav": "of which Buys (H) / of which Saveback (J)",
        "inv_totals_value": "current value of Buy shares (H) / Saveback shares (J)",
        # Yearly
        "yr_title": "Yearly summary — contributions and performance",
        "yr_cols": ["Year", "Buys (€)", "Saveback (€)", "Fees (€)", "Total cost (€)",
                    "Current value (€)", "Gain/loss (€)", "Gain/loss (%)"],
        # By ETF
        "etf_title": "Summary by ETF / stock",
        "etf_cols": ["Name / ETF", "Cumulative shares", "Buys (€)", "Saveback (€)",
                     "Total cost (€)", "Current price (€)", "Current value (€)",
                     "Gain/loss (€)", "Gain/loss (%)"],
        "etf_value_note": "\"Current value\" = cumulative shares × current price. "
                          "Gains are green, losses red.",
        "etf_global": "Global view — all ETFs",
        "etf_chart_gl_bar": "Gain/loss by ETF (global)",
        "etf_chart_gl_pie": "Value split (global)",
        "etf_year_section": "Per-year detail — pick a year in the dropdown",
        "etf_year_label": "Year:",
        "etf_by_line": "By ETF / stock",
        "etf_ycols": ["ETF / stock", "Shares bought", "Cost (€)", "Current value (€)",
                      "Gain/loss (€)", "Gain/loss (%)"],
        "etf_chart_y_bar": "Gain/loss by ETF (selected year)",
        "etf_chart_y_pie": "Value split (selected year)",
        "etf_month_section": "Monthly detail (selected year)",
        "etf_mcols": ["Month", "Buys (€)", "Saveback (€)", "Cost (€)",
                      "Value today (€)", "Gain/loss (€)"],
        "etf_chart_m_bar": "Gain/loss by month (selected year)",
        # Dashboard
        "db_title": "Dashboard — overview",
        "db_cards": ["Contributions\n(buys)", "Saveback\nreceived",
                     "Total cost\n(buys + saveback)", "Fees",
                     "Portfolio\ncurrent value", "Gain/loss (€)", "Gain/loss (%)"],
        "db_perf": "Performance",
        "db_perf_cols": ["—", "Basis", "Gain/loss (€)", "Gain/loss (%)", "What it measures"],
        "db_perf_tr": "Performance\n(TR method)",
        "db_perf_tr_m": "Return on all invested money",
        "db_perf_sav": "Saveback contribution",
        "db_perf_sav_m": "Today's value of offered shares (+ their perf)",
        "db_perf_note": "\"Performance\" = your market gain (like TR). "
                        "\"Saveback contribution\" = today's value of offered shares.",
        "db_month_section": "Monthly detail — portfolio value over time",
        "db_mcols": ["Month", "Buys (€)", "Saveback (€)", "Cost of month (€)",
                     "Cumulative cost (€)", "Portfolio value (€)",
                     "Unrealised gain/loss (€)", "Unrealised gain/loss (%)"],
        "db_month_note": "Portfolio value = your shares priced at each month's "
                         "transactions; last point = current price. Indicative "
                         "(prices at transaction dates).",
        "db_chart_line": "Invested cost vs portfolio value",
        "db_chart_bar": "Unrealised gain/loss over time (€)",
        "db_chart_pie": "Split (current value)",
        # Tax
        "tax_title": "Tax — summary for your filing",
        "tax_sub": "Ordinary securities account — flat tax 30% unless you opt for "
                   "the progressive scale.",
        "tax_cols": ["Year", "Cash interest\nreceived (€)", "Realised\ngains (€)",
                     "Taxable base\n(€)", "Estimated\ntax 30% (€)",
                     "Contributions\nof the year (€)"],
        "tax_notes": [
            "Cash interest: taxable in the year received (investment income).",
            "Accumulating ETFs (Acc): no dividend paid, taxed only on a sale "
            "(realised gain).",
            "Realised gains use the weighted-average cost method. Unrealised "
            "gains are not taxable.",
            "A tax-advantaged plan (e.g. PEA) would follow other rules — not "
            "covered here.",
            "Indicative estimate, not tax advice.",
        ],
        # Stock picking
        "sp_title": "Stock Picking & IPO — single stocks",
        "sp_sub": "Buys/sells outside ETFs. Cost basis at weighted-average price; "
                  "realised gain on sells. ISIN auto-filled, you only type the price.",
        "sp_prices": "Current prices",
        "sp_price_cols": ["Company", "ISIN", "Current price (€)", "Updated"],
        "sp_journal": "Transactions",
        "sp_jcols": ["Date", "Type", "Company", "ISIN", "Qty", "Price (€)",
                     "Fees (€)", "Amount (€)", "Comment"],
        "sp_empty": "No stock-picking transaction yet.",
        "sp_summary": "Summary by stock",
        "sp_scols": ["Company", "Shares held", "Avg cost (€)", "Remaining cost (€)",
                     "Current price (€)", "Current value (€)", "Unrealised g/l (€)",
                     "Unrealised g/l (%)", "Realised g/l (€)"],
        # CLI / console
        "cli_desc": "Build the Trade Republic board workbook from a CSV export.",
        "cli_fetch": "Fetching current prices (Deutsche Börse / Xetra, Yahoo fallback)…",
        "cli_price_line": "  price %s = %.2f €",
        "cli_missing": "  ! prices not found (fill by hand): %s",
        "cli_done": "Done -> %s",
        "cli_journal": "  ETF journal: %d rows (%d buys, %d savebacks, %d sells)",
        "cli_etf": "  ETFs: %s",
        "cli_stock": "  Stock picking: %d trades over %d stock(s)",
        "cli_tax": "  %s: interest %.2f € | realised gain %.2f € | contributions %.2f €",
        "cli_backup": "  (backup: %s.bak)",
        "cli_no_csv": "CSV not found: %s",
        "cli_bad_ext": "input must be a .csv file: %s",
        "cli_too_big": "file too large (> %d MB): %s",
        "cli_not_tr": "not a Trade Republic export (missing columns: %s)",
        "cli_extra_cols": "  ! unexpected extra columns ignored: %s",
        "cli_too_many_rows": "too many rows (> %d)",
        "cli_field_too_long": "a field exceeds %d characters (row %d)",
        "cli_bad_encoding": "file is not valid UTF-8: %s",
    },
    "fr": {
        "sheet_readme": "Lisez-moi",
        "sheet_dashboard": "Dashboard",
        "sheet_investments": "Investissements",
        "sheet_yearly": "Récap par année",
        "sheet_by_etf": "Récap par ETF",
        "sheet_tax": "Fiscalité",
        "sheet_stock": "Stock Picking-IPO",
        "buy": "Achat",
        "saveback": "Saveback",
        "sell": "Vente",
        "saveback_card": "Saveback carte",
        "roundup_card": "Arrondi carte",
        "ok": "OK",
        "gap": "écart",
        "total": "TOTAL",
        "totals": "TOTAUX",
        "col_date": "Date", "col_year": "Année", "col_month": "Mois",
        "col_type": "Type", "col_name": "Nom / ETF", "col_unit_price": "Prix du titre (€)",
        "col_shares": "Nb de parts", "col_amount": "Total (€)", "col_control": "Contrôle",
        "col_current_value": "Valeur actuelle (€)", "col_fees": "Frais (€)",
        "col_comment": "Commentaire", "col_isin": "ISIN", "col_current_price": "Cours actuel (€)",
        "col_updated": "Maj le", "col_etf_stock": "ETF / action",
        "rm_title": "Suivi d'investissements — Trade Republic",
        "rm_blocks": [
            ("Le principe", [
                "Chaque part que tu possèdes = une ligne du journal (Achat ou "
                "Saveback), avec son prix, ses parts, son total.",
                "Valeur du portefeuille = parts × cours du jour. Coût de revient = "
                "somme des totaux.",
                "Cellules JAUNES = à remplir (lignes saumon = savebacks). Le reste "
                "est automatique."]),
            ("Les frais", [
                "Le total d'une ligne = parts × prix (ce que TR investit "
                "réellement). Les frais d'ordre ponctuels vont dans la colonne "
                "Frais, séparément."]),
            ("Performance et couleurs", [
                "\"Performance\" (Dashboard) = ta plus-value de marché sur tout "
                "l'argent investi, comme l'affiche TR.",
                "\"Apport du saveback\" = ce que valent aujourd'hui les parts "
                "offertes par les savebacks.",
                "Toutes les +/- values sont en vert (gain) ou rouge (perte)."]),
            ("Les onglets", [
                "\"Investissements\" : cours du jour (ISIN auto) + journal ETF. "
                "Les écarts de contrôle passent en orange.",
                "\"Récap par année\" : versements et +/- value par année.",
                "\"Récap par ETF\" : performance par ligne, PLUS un détail par "
                "année à choisir dans un menu déroulant (avec graphiques).",
                "\"Dashboard\" : synthèse, +/- value, et un détail mensuel qui "
                "retrace la valeur du portefeuille au fil du temps (chaque mois "
                "valorisé au cours de ses transactions, dernier point = cours du jour).",
                "\"Fiscalité\" : intérêts espèce et plus-values réalisées "
                "(déclaration).",
                "\"Stock Picking / IPO\" : actions individuelles (achats/ventes, "
                "prix moyen pondéré)."]),
            ("Régénération mensuelle", [
                "Ré-exporte le CSV depuis Trade Republic puis relance le script : "
                "le journal est reconstruit et tes cours sont conservés.",
                "\"Cours automatiques\" : lance avec --auto-prices pour récupérer "
                "les cours par ISIN (Deutsche Börse / Xetra, repli Yahoo) ; sinon "
                "tu ne remplis que la colonne Cours actuel (jaune).",
                "Outil de suivi personnel, pas un conseil en investissement."]),
        ],
        "inv_title": "Investissements",
        "inv_prices_note": "Cours du jour — un cours par ETF (source : ton app TR). "
                           "ISIN rempli automatiquement, tu ne saisis que le cours.",
        "inv_journal_note": "Journal — jaune à remplir (saumon = saveback). "
                            "Total = parts × prix ; les frais vont dans leur colonne. "
                            "Ligne orange = écart de contrôle à corriger.",
        "inv_totals_buy_sav": "dont Achats (H) / dont Saveback (J)",
        "inv_totals_value": "valeur actuelle des parts Achats (H) / parts Saveback (J)",
        "yr_title": "Récapitulatif par année — versements et performance",
        "yr_cols": ["Année", "Achats (€)", "Saveback (€)", "Frais (€)", "Coût total (€)",
                    "Valeur actuelle (€)", "+/- value (€)", "+/- value (%)"],
        "etf_title": "Récapitulatif par ETF / action",
        "etf_cols": ["Nom / ETF", "Parts cumulées", "Achats (€)", "Saveback (€)",
                     "Coût total (€)", "Cours du jour (€)", "Valeur actuelle (€)",
                     "+/- value (€)", "+/- value (%)"],
        "etf_value_note": "\"Valeur actuelle\" = parts cumulées × cours du jour. "
                          "Les +/- values sont en vert (gain) ou rouge (perte).",
        "etf_global": "Vue globale — tous les ETF",
        "etf_chart_gl_bar": "+/- value par ETF (global)",
        "etf_chart_gl_pie": "Répartition valeur (global)",
        "etf_year_section": "Détail par année — choisis l'année dans le menu déroulant",
        "etf_year_label": "Année :",
        "etf_by_line": "Par ETF / action",
        "etf_ycols": ["ETF / action", "Parts acquises", "Coût (€)", "Valeur actuelle (€)",
                      "+/- value (€)", "+/- value (%)"],
        "etf_chart_y_bar": "+/- value par ETF (année choisie)",
        "etf_chart_y_pie": "Répartition valeur (année choisie)",
        "etf_month_section": "Détail mensuel (année choisie)",
        "etf_mcols": ["Mois", "Achats (€)", "Saveback (€)", "Coût (€)",
                      "Valeur auj. (€)", "+/- value (€)"],
        "etf_chart_m_bar": "+/- value par mois (année choisie)",
        "db_title": "Dashboard — synthèse globale",
        "db_cards": ["Versements\n(achats)", "Saveback\nreçu",
                     "Coût total\n(achats + saveback)", "Frais",
                     "Valeur actuelle\nportefeuille", "+/- value (€)", "+/- value (%)"],
        "db_perf": "Performance",
        "db_perf_cols": ["—", "Base de calcul", "+/- value (€)", "+/- value (%)",
                         "Ce que ça mesure"],
        "db_perf_tr": "Performance\n(méthode TR)",
        "db_perf_tr_m": "Rendement de tout l'argent investi",
        "db_perf_sav": "Apport du saveback",
        "db_perf_sav_m": "Valeur actuelle des parts offertes (+ leur perf)",
        "db_perf_note": "\"Performance\" = ta plus-value de marché (comme l'affiche "
                        "TR). \"Apport du saveback\" = ce que valent aujourd'hui les "
                        "parts offertes.",
        "db_month_section": "Détail mensuel — valeur du portefeuille au fil du temps",
        "db_mcols": ["Mois", "Achats (€)", "Saveback (€)", "Coût du mois (€)",
                     "Cumul coût (€)", "Valeur portefeuille (€)",
                     "+/- value latente (€)", "+/- value latente (%)"],
        "db_month_note": "Valeur portefeuille = tes parts valorisées au cours de "
                         "chaque mois (prix des transactions TR) ; dernier point = "
                         "cours du jour. Vue indicative (cours en date de transaction).",
        "db_chart_line": "Coût investi vs valeur du portefeuille",
        "db_chart_bar": "+/- value latente au fil du temps (€)",
        "db_chart_pie": "Répartition (valeur actuelle)",
        "tax_title": "Fiscalité — récapitulatif pour la déclaration",
        "tax_sub": "Compte-titres ordinaire (CTO) — imposition au PFU 30 % "
                   "(flat tax) sauf option barème.",
        "tax_cols": ["Année", "Intérêts espèces\nperçus (€)", "Plus-values\nréalisées (€)",
                     "Base imposable\n(€)", "Estimation\nprélèvement 30 % (€)",
                     "Versements\nde l'année (€)"],
        "tax_notes": [
            "Intérêts espèces : imposables l'année de perception (revenus de "
            "capitaux mobiliers, case 2TR).",
            "ETF capitalisants (Acc) : aucun dividende distribué, imposition "
            "uniquement lors d'une revente (plus-value réalisée).",
            "Plus-values réalisées calculées au prix moyen pondéré. La +/- value "
            "latente n'est pas imposable.",
            "Un PEA suivrait d'autres règles (exonération après 5 ans) — non "
            "concerné ici.",
            "Estimation indicative, ce n'est pas un conseil fiscal.",
        ],
        "sp_title": "Stock Picking & IPO — actions individuelles",
        "sp_sub": "Achats/ventes hors ETF. Prix de revient au prix moyen pondéré ; "
                  "+/- value réalisée sur les ventes. ISIN rempli automatiquement, "
                  "tu ne saisis que le cours.",
        "sp_prices": "Cours du jour",
        "sp_price_cols": ["Société", "ISIN", "Cours actuel (€)", "Maj le"],
        "sp_journal": "Journal des opérations",
        "sp_jcols": ["Date", "Type", "Société", "ISIN", "Qté", "Prix (€)",
                     "Frais (€)", "Montant (€)", "Commentaire"],
        "sp_empty": "Aucune opération de stock picking pour l'instant.",
        "sp_summary": "Synthèse par titre",
        "sp_scols": ["Société", "Qté détenue", "PMP (€)", "Coût restant (€)",
                     "Cours du jour (€)", "Valeur actuelle (€)", "+/- value latente (€)",
                     "+/- value latente (%)", "+/- value réalisée (€)"],
        "cli_desc": "Génère le classeur de suivi Trade Republic depuis le CSV.",
        "cli_fetch": "Récupération des cours (Deutsche Börse / Xetra, repli Yahoo)…",
        "cli_price_line": "  cours %s = %.2f €",
        "cli_missing": "  ! cours introuvables (à saisir à la main) : %s",
        "cli_done": "OK -> %s",
        "cli_journal": "  Journal ETF : %d lignes (%d achats, %d savebacks, %d ventes)",
        "cli_etf": "  ETF : %s",
        "cli_stock": "  Stock picking : %d opérations sur %d titre(s)",
        "cli_tax": "  %s : intérêts %.2f € | PV réalisée %.2f € | versements %.2f €",
        "cli_backup": "  (sauvegarde : %s.bak)",
        "cli_no_csv": "CSV introuvable : %s",
        "cli_bad_ext": "le fichier d'entrée doit être un .csv : %s",
        "cli_too_big": "fichier trop volumineux (> %d Mo) : %s",
        "cli_not_tr": "pas un export Trade Republic (colonnes manquantes : %s)",
        "cli_extra_cols": "  ! colonnes supplémentaires inattendues ignorées : %s",
        "cli_too_many_rows": "trop de lignes (> %d)",
        "cli_field_too_long": "un champ dépasse %d caractères (ligne %d)",
        "cli_bad_encoding": "fichier non valide en UTF-8 : %s",
    },
}

# Active-language globals, set by select_language() before any building.
LANG = "en"
S = TXT["en"]
MONTHS = MONTH_NAMES["en"]
MONTHS_ABBR = MONTH_ABBR["en"]
INV = TXT["en"]["sheet_investments"]           # investments sheet name (used in formulas)
_ISIN_NAMES = {}                               # isin -> display name, filled from the CSV


def select_language(lang):
    global LANG, S, MONTHS, MONTHS_ABBR, INV
    LANG = lang if lang in TXT else "en"
    S = TXT[LANG]
    MONTHS = MONTH_NAMES[LANG]
    MONTHS_ABBR = MONTH_ABBR[LANG]
    INV = S["sheet_investments"]


# ==========================================================================
#  Palette and styles
# ==========================================================================

def C(hexa):                                   # -> ARGB colour
    return "FF" + hexa

NAVY, BLUE, LIGHT_BLUE = "1F3A5F", "2E5E8C", "EAF1F8"
YELLOW, SALMON, GOLD = "FFF2CC", "FDECE4", "C9A227"
WHITE, ORANGE = "FFFFFF", "F8CBAD"
C_INPUT, C_CALC, C_CTRL = "0000FF", "000000", "008000"
GREY1, GREY2, GREY3 = "888888", "666666", "333333"

FMT_EUR = '#,##0.00" €";[RED]\\-#,##0.00" €"'
# Gains/losses are coloured by conditional formatting (see colour_pl), not by the
# number format, so the format stays neutral here.
FMT_EUR_PL = '#,##0.00" €";\\-#,##0.00" €"'
FMT_PCT_PL = '0.0%;\\-0.0%'
GREEN_PL = "FF1B7F3B"                           # dark green, readable on white
RED_PL = "FFC0392B"                             # dark red
FMT_PRICE = '#,##0.0000" €"'
FMT_SHARES = "0.000000"
FMT_DATE = "dd/mm/yyyy"

F_TITLE = Font(bold=True, size=16, color=C(NAVY))
F_TITLE_XL = Font(bold=True, size=18, color=C(NAVY))
F_TITLE_RM = Font(bold=True, size=20, color=C(NAVY))
F_SECTION = Font(bold=True, size=11, color=C(WHITE))
F_HDR = Font(bold=True, size=10, color=C(WHITE))
F_HDR9 = Font(bold=True, size=9, color=C(WHITE))
F_LABEL = Font(bold=True, size=10, color=C(NAVY))
F_INPUT = Font(size=10, color=C(C_INPUT))
F_CALC = Font(size=10, color=C(C_CALC))
F_CTRL = Font(size=10, color=C(C_CTRL))
F_TOTAL = Font(bold=True, size=10, color=C(NAVY))
F_KPI = Font(bold=True, size=13, color=C(NAVY))
F_NOTE = Font(size=9, color=C(GREY2))
F_SUB = Font(size=9, italic=True, color=C(GREY1))
F_TXT = Font(size=10, color=C(GREY3))
F_GOLD = Font(bold=True, size=10, color=C(WHITE))


def fill(hexa):
    return PatternFill("solid", fgColor=C(hexa))

FILL_NAVY, FILL_BLUE, FILL_LIGHT = fill(NAVY), fill(BLUE), fill(LIGHT_BLUE)
FILL_YELLOW, FILL_SALMON, FILL_GOLD = fill(YELLOW), fill(SALMON), fill(GOLD)
FILL_WHITE = fill(WHITE)

A_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
A_CV = Alignment(horizontal="center", vertical="center")
A_L = Alignment(horizontal="left", vertical="center")
A_LW = Alignment(horizontal="left", vertical="center", wrap_text=True)

_side = Side(style="thin", color="FFBFBFBF")
B_ALL = Border(left=_side, right=_side, top=_side, bottom=_side)


def put(ws, r, c, value=None, font=F_CALC, fl=None, fmt=None, align=A_CV, border=None):
    """Write a cell and apply font/fill/format/alignment/border in one call."""
    cell = ws.cell(r, c, value)
    cell.font = font
    if fl is not None:
        cell.fill = fl
    if fmt is not None:
        cell.number_format = fmt
    if align is not None:
        cell.alignment = align
    if border is not None:
        cell.border = border
    return cell


def widths(ws, mapping):
    for col, w in mapping.items():
        ws.column_dimensions[col].width = w


def colour_pl(ws, *ranges):
    """Colour gain/loss cells green (>0) / red (<0) via conditional formatting."""
    for rng in ranges:
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="greaterThan", formula=["0"], font=Font(color=GREEN_PL)))
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="lessThan", formula=["0"], font=Font(color=RED_PL)))


# ==========================================================================
#  Data layer
# ==========================================================================

# --- Hardened ingestion (treat the CSV as untrusted input) -----------------
#
# The columns the tool actually reads. Their presence is the signature of a
# Trade Republic export; extra columns are tolerated (survives TR schema tweaks).
REQUIRED_COLUMNS = ["date", "category", "type", "asset_class", "name", "symbol",
                    "shares", "price", "amount", "fee"]
# The full known Trade Republic export schema — used only to decide whether an
# extra column is genuinely unexpected (worth a warning) or just an unused field.
KNOWN_COLUMNS = set(REQUIRED_COLUMNS) | {
    "datetime", "account_type", "tax", "currency", "original_amount",
    "original_currency", "fx_rate", "description", "transaction_id",
    "counterparty_name", "counterparty_iban", "payment_reference", "mcc_code"}
MAX_FILE_BYTES = 10 * 1024 * 1024                  # 10 MB — TR exports are tiny
MAX_ROWS = 200_000
MAX_FIELD_LEN = 2000


def safe_text(s):
    """Neutralise CSV / formula injection for any CSV-derived string written to a
    cell. openpyxl treats a value starting with '=' as a formula; '+ - @' are the
    other classic spreadsheet-injection triggers. Prefix them so Excel keeps text.
    """
    if isinstance(s, str) and s[:1] in ("=", "+", "-", "@"):
        return "'" + s
    return s


def validate_export(path):
    """Validate an untrusted CSV and return its rows, or raise SystemExit.

    Checks: extension, size cap, UTF-8 decoding, required columns, row and field
    caps. Reused by the CLI and the watcher.
    """
    if not os.path.exists(path):
        raise SystemExit(S["cli_no_csv"] % path)
    if not path.lower().endswith(".csv"):
        raise SystemExit(S["cli_bad_ext"] % path)
    if os.path.getsize(path) > MAX_FILE_BYTES:
        raise SystemExit(S["cli_too_big"] % (MAX_FILE_BYTES // (1024 * 1024), path))
    try:
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            fields = reader.fieldnames or []
            missing = [c for c in REQUIRED_COLUMNS if c not in fields]
            if missing:
                raise SystemExit(S["cli_not_tr"] % ", ".join(missing))
            extra = [c for c in fields if c and c not in KNOWN_COLUMNS]
            if extra:
                print(S["cli_extra_cols"] % ", ".join(extra))
            rows = []
            for i, row in enumerate(reader):
                if i >= MAX_ROWS:
                    raise SystemExit(S["cli_too_many_rows"] % MAX_ROWS)
                for v in row.values():
                    if isinstance(v, str) and len(v) > MAX_FIELD_LEN:
                        raise SystemExit(S["cli_field_too_long"] % (MAX_FIELD_LEN, i + 2))
                rows.append(row)
    except UnicodeDecodeError:
        raise SystemExit(S["cli_bad_encoding"] % path)
    return rows


def num(val):
    val = (val or "").strip()
    return float(val) if val else None


def index_names(rows):
    """Build isin -> display name (most frequent) from the CSV."""
    counts = defaultdict(Counter)
    for r in rows:
        if r["symbol"] and r["name"]:
            counts[r["symbol"]][r["name"]] += 1
    return {isin: c.most_common(1)[0][0] for isin, c in counts.items()}


def display_name(row):
    return safe_text(_ISIN_NAMES.get(row["symbol"]) or row["name"] or row["symbol"] or "?")


def norm_name(s):
    """Matching key for names, insensitive to a currency token."""
    s = str(s or "")
    for tok in (" EUR ", " USD ", " GBP ", " CHF "):
        s = s.replace(tok, " ")
    return " ".join(s.split()).lower()


def is_trade(r):
    return r["category"] == "TRADING" or r["type"] in ("BUY", "SELL")


def direction(r):
    """Return 'BUY' or 'SELL' from the type / the sign of the amount."""
    if "SELL" in r["type"] or "SALE" in r["type"]:
        return "SELL"
    if r["type"] == "BUY":
        return "BUY"
    a = num(r["amount"]) or 0
    return "SELL" if a > 0 else "BUY"


def to_trade(r):
    return {
        "date": datetime.strptime(r["date"], "%Y-%m-%d"),
        "dir": direction(r),
        "isin": r["symbol"],
        "name": display_name(r),
        "qty": abs(num(r["shares"]) or 0),
        "price": num(r["price"]) or 0,
        "amount": abs(num(r["amount"]) or 0),
        "fee": abs(num(r["fee"]) or 0),
    }


def is_roundup(benefit_type):
    """Trade Republic "Round up" / spare-change benefit (vs saveback)."""
    return "SPARE" in benefit_type or "ROUND" in benefit_type


def build_etf_journal(rows):
    """Ordered ETF journal entries (Buy / Saveback / Sell).

    A benefit credit (saveback or round-up: any BENEFITS_* type) reinvested by a
    buy of the same amount tags that buy as a "Saveback" (matching by amount +
    closest date). Matching on any BENEFITS_* is safe -- a credit with no
    reinvesting buy is simply ignored -- and future-proofs round-up handling.
    """
    fund = [r for r in rows if is_trade(r) and r["asset_class"] == "FUND"]
    buys = [r for r in fund if direction(r) == "BUY"]
    sells = [r for r in fund if direction(r) == "SELL"]
    benefits = [r for r in rows if r["type"].startswith("BENEFITS_")]

    used = [False] * len(buys)
    benefit_of = {}                                # buy index -> benefit credit type
    for cr in benefits:
        m = round(abs(num(cr["amount"]) or 0), 2)
        d = datetime.strptime(cr["date"], "%Y-%m-%d")
        best, bkey = None, None
        for i, b in enumerate(buys):
            if used[i] or round(abs(num(b["amount"]) or 0), 2) != m:
                continue
            db = datetime.strptime(b["date"], "%Y-%m-%d")
            key = (0 if db >= d else 1, abs((db - d).days))
            if bkey is None or key < bkey:
                bkey, best = key, i
        if best is not None:
            used[best] = True
            benefit_of[best] = cr["type"]

    entries = []
    for i, b in enumerate(buys):
        benefit = i in benefit_of
        comment = None
        if benefit:
            comment = S["roundup_card"] if is_roundup(benefit_of[i]) else S["saveback_card"]
        entries.append(dict(
            date=datetime.strptime(b["date"], "%Y-%m-%d"),
            kind="saveback" if benefit else "buy",
            type=S["saveback"] if benefit else S["buy"],
            name=display_name(b), price=num(b["price"]), shares=num(b["shares"]),
            total=round(abs(num(b["amount"]) or 0), 2),
            fee=round(abs(num(b["fee"]) or 0), 2),
            comment=comment))
    for v in sells:
        entries.append(dict(
            date=datetime.strptime(v["date"], "%Y-%m-%d"), kind="sell", type=S["sell"],
            name=display_name(v), price=num(v["price"]),
            shares=-abs(num(v["shares"]) or 0),
            total=-round(abs(num(v["amount"]) or 0), 2),
            fee=round(abs(num(v["fee"]) or 0), 2), comment=S["sell"]))

    rank = {"buy": 0, "sell": 0, "saveback": 1}
    entries.sort(key=lambda e: (e["date"], rank.get(e["kind"], 2), e["name"]))
    return entries


def positions_wac(trades):
    """Weighted-average cost: current state per stock + realised gain per year/isin."""
    state, real_year, real_isin = {}, defaultdict(float), defaultdict(float)
    for t in sorted(trades, key=lambda x: x["date"]):
        s = state.setdefault(t["isin"], dict(qty=0.0, cost=0.0, name=t["name"],
                                             isin=t["isin"]))
        if t["dir"] == "BUY":
            s["qty"] += t["qty"]
            s["cost"] += t["qty"] * t["price"] + t["fee"]
        else:
            wac = s["cost"] / s["qty"] if s["qty"] else 0
            gain = t["qty"] * t["price"] - t["qty"] * wac - t["fee"]
            real_year[t["date"].year] += gain
            real_isin[t["isin"]] += gain
            s["cost"] -= t["qty"] * wac
            s["qty"] -= t["qty"]
    return state, real_year, real_isin


def stock_trades(rows):
    return [to_trade(r) for r in rows
            if is_trade(r) and r["asset_class"] != "FUND" and r["symbol"]]


def tax_summary(rows, entries, real_year):
    """Per year: taxable cash interest, realised gains, contributions."""
    years = {}

    def slot(a):
        return years.setdefault(str(a), dict(interest=0.0, gain=0.0, contrib=0.0))

    for r in rows:
        if r["type"] == "INTEREST_PAYMENT":
            slot(r["date"][:4])["interest"] += num(r["amount"]) or 0
    for e in entries:
        if e["kind"] != "sell":
            slot(e["date"].year)["contrib"] += e["total"]
    for y, g in real_year.items():
        slot(y)["gain"] += g
    return dict(sorted(years.items()))


def covered_months(entries):
    if not entries:
        return []
    lo = min(e["date"] for e in entries)
    hi = max(e["date"] for e in entries)
    out, y, m = [], lo.year, lo.month
    while (y, m) <= (hi.year, hi.month):
        out.append((y, m))
        m += 1
        if m > 12:
            m, y = 1, y + 1
    return out


def monthly_historic_value(entries, months):
    """Mark-to-market portfolio value at the end of each month, using the prices
    of the Trade Republic transactions.

    For each month: cumulative shares held × that month's price. A stock's price
    is that of its last transaction in the month; if it is held but not traded
    that month, the last known price is carried forward (it was necessarily
    bought earlier). Negative shares (sells) reduce the running total.
    """
    month_price, month_shares = {}, defaultdict(float)
    names = []
    for e in entries:
        key = (e["name"], (e["date"].year, e["date"].month))
        if e["price"]:
            month_price[key] = e["price"]
        month_shares[key] += (e["shares"] or 0)
        if e["name"] not in names:
            names.append(e["name"])

    cum, last_price, values = defaultdict(float), {}, []
    for ym in months:
        for name in names:
            if (name, ym) in month_price:
                last_price[name] = month_price[(name, ym)]
            cum[name] += month_shares.get((name, ym), 0.0)
        v = sum(cum[name] * last_price[name] for name in names if name in last_price)
        values.append(round(v, 2))
    return values


# ==========================================================================
#  Read me sheet
# ==========================================================================

def build_readme(ws):
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 3, "B": 112})
    put(ws, 2, 2, S["rm_title"], F_TITLE_RM, align=A_L)
    r = 4
    for title, lines in S["rm_blocks"]:
        put(ws, r, 2, title, F_SECTION, FILL_BLUE, align=A_L)
        r += 1
        for line in lines:
            put(ws, r, 2, line, F_TXT, align=A_LW)
            ws.row_dimensions[r].height = 15
            r += 1
        r += 1


# ==========================================================================
#  Investments sheet (ETF journal) -> returns the layout
# ==========================================================================

def build_investments(ws, entries, name_isin, by_isin, by_name):
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 26.6, "B": 14.2, "C": 13, "D": 12, "E": 26, "F": 15,
                "G": 13, "H": 14, "I": 11, "J": 16, "K": 10, "L": 24})

    put(ws, 1, 1, S["inv_title"], F_TITLE, align=A_L)
    ws.merge_cells("A1:L1")

    # --- Current prices (ISIN auto; you only type the price) ---
    put(ws, 3, 1, S["inv_prices_note"], F_SECTION, FILL_BLUE, align=A_L)
    ws.merge_cells("A3:D3")
    for j, txt in enumerate([S["col_etf_stock"], S["col_isin"],
                             S["col_current_price"], S["col_updated"]]):
        put(ws, 4, 1 + j, txt, F_HDR, FILL_NAVY, align=A_C, border=B_ALL)

    names = []
    for e in entries:
        if e["name"] not in names:
            names.append(e["name"])
    n_slots = max(len(names) + 3, 5)
    ETF_FIRST = 5
    ETF_LAST = ETF_FIRST + n_slots - 1
    for i in range(n_slots):
        r = ETF_FIRST + i
        name = names[i] if i < len(names) else None
        isin = name_isin.get(name) if name else None
        price, updated = price_for(isin, name, by_isin, by_name) if name else (None, None)
        put(ws, r, 1, name, F_LABEL, FILL_LIGHT, align=A_L, border=B_ALL)
        put(ws, r, 2, safe_text(isin), F_CALC, FILL_LIGHT, None, A_CV, B_ALL)
        put(ws, r, 3, price, F_INPUT, FILL_YELLOW, FMT_PRICE, border=B_ALL)
        put(ws, r, 4, updated, F_INPUT, FILL_YELLOW, FMT_DATE, border=B_ALL)
    etf_range = "$A${a}:$C${b}".format(a=ETF_FIRST, b=ETF_LAST)

    # --- Journal ---
    NOTE = ETF_LAST + 2
    HDR = NOTE + 1
    FIRST = HDR + 1
    N_ROWS = 320
    LAST = FIRST + N_ROWS - 1

    put(ws, NOTE, 1, S["inv_journal_note"], F_SUB, align=A_L)
    ws.merge_cells(start_row=NOTE, start_column=1, end_row=NOTE, end_column=12)

    headers = [S["col_date"], S["col_year"], S["col_month"], S["col_type"],
               S["col_name"], S["col_unit_price"], S["col_shares"], S["col_amount"],
               S["col_control"], S["col_current_value"], S["col_fees"], S["col_comment"]]
    for j, txt in enumerate(headers):
        put(ws, HDR, 1 + j, txt, F_HDR, FILL_NAVY, align=A_C, border=B_ALL)

    months_formula = ",".join('"%s"' % m for m in MONTHS)
    gap, ok = S["gap"], S["ok"]
    for k in range(FIRST, LAST + 1):
        e = entries[k - FIRST] if (k - FIRST) < len(entries) else None
        is_saveback = bool(e) and e["kind"] == "saveback"
        bg = FILL_SALMON if is_saveback else FILL_YELLOW
        # editable columns
        put(ws, k, 1, e["date"] if e else None, F_INPUT, bg, FMT_DATE, border=B_ALL)
        put(ws, k, 4, e["type"] if e else None, F_INPUT, bg, align=A_CV, border=B_ALL)
        put(ws, k, 5, e["name"] if e else None, F_INPUT, bg, align=A_L, border=B_ALL)
        put(ws, k, 6, e["price"] if e else None, F_INPUT, bg, FMT_PRICE, border=B_ALL)
        put(ws, k, 7, e["shares"] if e else None, F_INPUT, bg, FMT_SHARES, border=B_ALL)
        put(ws, k, 8, e["total"] if e else None, F_INPUT, bg, FMT_EUR, border=B_ALL)
        put(ws, k, 11, e["fee"] if e else None, F_INPUT, bg, FMT_EUR, border=B_ALL)
        put(ws, k, 12, e["comment"] if e else None, F_INPUT, bg, align=A_L, border=B_ALL)
        # computed columns
        put(ws, k, 2, '=IF($A{k}="","",YEAR($A{k}))'.format(k=k),
            F_CALC, None, None, A_CV, B_ALL)
        put(ws, k, 3, '=IF($A{k}="","",CHOOSE(MONTH($A{k}),{m}))'.format(k=k, m=months_formula),
            F_CALC, None, None, A_CV, B_ALL)
        put(ws, k, 9, ('=IF(OR($F{k}="",$G{k}="",$H{k}=""),"",'
                       'IF(ABS($F{k}*$G{k}-$H{k})<=0.05,"{ok}",'
                       '"{gap} "&TEXT($F{k}*$G{k}-$H{k},"0.00")))').format(k=k, ok=ok, gap=gap),
            F_CTRL, None, None, A_CV, B_ALL)
        put(ws, k, 10, ('=IF(OR($G{k}="",IFERROR(VLOOKUP($E{k},{rg},3,0),"")=""),"",'
                        '$G{k}*IFERROR(VLOOKUP($E{k},{rg},3,0),""))').format(k=k, rg=etf_range),
            F_CALC, None, FMT_EUR, A_CV, B_ALL)

    # --- Totals ---
    TOT = LAST + 2
    put(ws, TOT, 1, S["totals"], F_TOTAL, align=A_L)
    ws.merge_cells(start_row=TOT, start_column=1, end_row=TOT, end_column=5)
    put(ws, TOT, 8, "=SUM(H{a}:H{b})".format(a=FIRST, b=LAST), F_TOTAL, None, FMT_EUR)
    put(ws, TOT, 10, "=SUM(J{a}:J{b})".format(a=FIRST, b=LAST), F_TOTAL, None, FMT_EUR)
    put(ws, TOT, 11, "=SUM(K{a}:K{b})".format(a=FIRST, b=LAST), F_TOTAL, None, FMT_EUR)

    put(ws, TOT + 1, 1, S["inv_totals_buy_sav"], F_NOTE, align=A_L)
    ws.merge_cells(start_row=TOT + 1, start_column=1, end_row=TOT + 1, end_column=5)
    put(ws, TOT + 1, 8, '=SUMIF($D${a}:$D${b},"{buy}",$H${a}:$H${b})'.format(
        a=FIRST, b=LAST, buy=S["buy"]), F_TOTAL, None, FMT_EUR)
    put(ws, TOT + 1, 10, '=SUMIF($D${a}:$D${b},"{sb}",$H${a}:$H${b})'.format(
        a=FIRST, b=LAST, sb=S["saveback"]), F_TOTAL, None, FMT_EUR)

    put(ws, TOT + 2, 1, S["inv_totals_value"], F_NOTE, align=A_L)
    ws.merge_cells(start_row=TOT + 2, start_column=1, end_row=TOT + 2, end_column=5)
    put(ws, TOT + 2, 8, '=SUMIF($D${a}:$D${b},"{buy}",$J${a}:$J${b})'.format(
        a=FIRST, b=LAST, buy=S["buy"]), F_TOTAL, None, FMT_EUR)
    put(ws, TOT + 2, 10, '=SUMIF($D${a}:$D${b},"{sb}",$J${a}:$J${b})'.format(
        a=FIRST, b=LAST, sb=S["saveback"]), F_TOTAL, None, FMT_EUR)

    # Conditional formatting: a consistency gap turns the row orange.
    ws.conditional_formatting.add(
        "A{a}:L{b}".format(a=FIRST, b=LAST),
        FormulaRule(formula=['LEFT($I{a},{n})="{gap}"'.format(a=FIRST, n=len(gap), gap=gap)],
                    fill=fill(ORANGE), stopIfTrue=False))

    ws.freeze_panes = "A{}".format(FIRST)
    return dict(etf_range=etf_range, etf_first=ETF_FIRST, etf_last=ETF_LAST,
                first=FIRST, last=LAST, tot=TOT, names=names)


# ==========================================================================
#  Yearly summary sheet
# ==========================================================================

def build_yearly(ws, L, years):
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 12, "B": 15, "C": 14, "D": 11, "E": 15, "F": 15, "G": 15, "H": 13})
    put(ws, 1, 1, S["yr_title"], F_TITLE, align=A_L)
    ws.merge_cells("A1:H1")
    for j, txt in enumerate(S["yr_cols"]):
        put(ws, 3, 1 + j, txt, F_HDR9, FILL_NAVY, align=A_C, border=B_ALL)

    a, b = L["first"], L["last"]
    buy, sb = S["buy"], S["saveback"]
    r0 = 4
    for i, yr in enumerate(years):
        r = r0 + i
        put(ws, r, 1, int(yr), F_LABEL, FILL_LIGHT, align=A_CV, border=B_ALL)
        put(ws, r, 2, '=SUMIFS({INV}!$H${a}:$H${b},{INV}!$B${a}:$B${b},$A{r},'
            '{INV}!$D${a}:$D${b},"{buy}")'.format(INV=INV, a=a, b=b, r=r, buy=buy),
            F_CALC, None, FMT_EUR, border=B_ALL)
        put(ws, r, 3, '=SUMIFS({INV}!$H${a}:$H${b},{INV}!$B${a}:$B${b},$A{r},'
            '{INV}!$D${a}:$D${b},"{sb}")'.format(INV=INV, a=a, b=b, r=r, sb=sb),
            F_CALC, None, FMT_EUR, border=B_ALL)
        put(ws, r, 4, '=SUMIF({INV}!$B${a}:$B${b},$A{r},{INV}!$K${a}:$K${b})'.format(
            INV=INV, a=a, b=b, r=r), F_CALC, None, FMT_EUR, border=B_ALL)
        put(ws, r, 5, "=B{r}+C{r}".format(r=r), F_CALC, None, FMT_EUR, border=B_ALL)
        put(ws, r, 6, '=SUMIFS({INV}!$J${a}:$J${b},{INV}!$B${a}:$B${b},$A{r})'.format(
            INV=INV, a=a, b=b, r=r), F_CALC, None, FMT_EUR, border=B_ALL)
        put(ws, r, 7, '=IF(F{r}=0,"",F{r}-E{r})'.format(r=r), F_CALC, None, FMT_EUR_PL, border=B_ALL)
        put(ws, r, 8, '=IF(OR(E{r}=0,F{r}=0),"",(F{r}-E{r})/E{r})'.format(r=r),
            F_CALC, None, FMT_PCT_PL, border=B_ALL)

    rt = r0 + len(years)
    put(ws, rt, 1, S["total"], F_TOTAL, align=A_CV, border=B_ALL)
    for col, f in [(2, FMT_EUR), (3, FMT_EUR), (4, FMT_EUR), (5, FMT_EUR),
                   (6, FMT_EUR), (7, FMT_EUR_PL)]:
        Lc = get_column_letter(col)
        put(ws, rt, col, "=SUM({L}{a}:{L}{b})".format(L=Lc, a=r0, b=rt - 1),
            F_TOTAL, None, f, border=B_ALL)
    put(ws, rt, 8, '=IF(OR(E{rt}=0,F{rt}=0),"",(F{rt}-E{rt})/E{rt})'.format(rt=rt),
        F_TOTAL, None, FMT_PCT_PL, border=B_ALL)
    colour_pl(ws, "G{a}:H{b}".format(a=r0, b=rt))


# ==========================================================================
#  By-ETF sheet
# ==========================================================================

def build_by_etf(ws, L, years):
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 30, "B": 14, "C": 15, "D": 13, "E": 15, "F": 14, "G": 16,
                "H": 15, "I": 13})
    put(ws, 1, 1, S["etf_title"], F_TITLE, align=A_L)
    ws.merge_cells("A1:I1")
    for j, txt in enumerate(S["etf_cols"]):
        put(ws, 3, 1 + j, txt, F_HDR9, FILL_NAVY, align=A_C, border=B_ALL)

    a, b = L["first"], L["last"]
    rg = L["etf_range"]
    names = L["names"]
    buy, sb = S["buy"], S["saveback"]
    n = max(len(names), 6)
    r0 = 4
    for i in range(n):
        r = r0 + i
        name = names[i] if i < len(names) else None
        put(ws, r, 1, name, F_LABEL, FILL_LIGHT, align=A_L, border=B_ALL)
        put(ws, r, 2, '=IF($A{r}="","",SUMIF({INV}!$E${a}:$E${b},$A{r},{INV}!$G${a}:$G${b}))'.format(
            INV=INV, a=a, b=b, r=r), F_CALC, None, FMT_SHARES, border=B_ALL)
        put(ws, r, 3, '=IF($A{r}="","",SUMIFS({INV}!$H${a}:$H${b},{INV}!$E${a}:$E${b},$A{r},'
            '{INV}!$D${a}:$D${b},"{buy}"))'.format(INV=INV, a=a, b=b, r=r, buy=buy),
            F_CALC, None, FMT_EUR, border=B_ALL)
        put(ws, r, 4, '=IF($A{r}="","",SUMIFS({INV}!$H${a}:$H${b},{INV}!$E${a}:$E${b},$A{r},'
            '{INV}!$D${a}:$D${b},"{sb}"))'.format(INV=INV, a=a, b=b, r=r, sb=sb),
            F_CALC, None, FMT_EUR, border=B_ALL)
        put(ws, r, 5, '=IF($A{r}="","",C{r}+D{r})'.format(r=r), F_CALC, None, FMT_EUR, border=B_ALL)
        put(ws, r, 6, '=IF($A{r}="","",IFERROR(VLOOKUP($A{r},{INV}!{rg},3,0),""))'.format(
            INV=INV, rg=rg, r=r), F_CALC, None, FMT_PRICE, border=B_ALL)
        put(ws, r, 7, '=IF(OR($A{r}="",B{r}="",F{r}=""),"",B{r}*F{r})'.format(r=r),
            F_CALC, None, FMT_EUR, border=B_ALL)
        put(ws, r, 8, '=IF(G{r}="","",G{r}-E{r})'.format(r=r), F_CALC, None, FMT_EUR_PL, border=B_ALL)
        put(ws, r, 9, '=IF(OR(E{r}=0,G{r}=""),"",H{r}/E{r})'.format(r=r),
            F_CALC, None, FMT_PCT_PL, border=B_ALL)

    rt = r0 + n
    put(ws, rt, 1, S["total"], F_TOTAL, align=A_L, border=B_ALL)
    for col, f in [(2, FMT_SHARES), (3, FMT_EUR), (4, FMT_EUR), (5, FMT_EUR),
                   (7, FMT_EUR), (8, FMT_EUR_PL)]:
        Lc = get_column_letter(col)
        put(ws, rt, col, "=SUM({L}{a}:{L}{b})".format(L=Lc, a=r0, b=rt - 1),
            F_TOTAL, None, f, border=B_ALL)
    put(ws, rt, 9, '=IF(E{rt}=0,"",H{rt}/E{rt})'.format(rt=rt), F_TOTAL, None, FMT_PCT_PL, border=B_ALL)
    colour_pl(ws, "H{a}:I{b}".format(a=r0, b=rt))

    put(ws, rt + 2, 1, S["etf_value_note"], F_NOTE, align=A_LW)
    ws.merge_cells(start_row=rt + 2, start_column=1, end_row=rt + 2, end_column=9)

    real_names = [x for x in names if x]
    nd = max(len(real_names), 1)
    ge = r0 + max(len(real_names), 1) - 1        # last real ETF row (main table)

    # --- Global view (all ETFs) ---
    GT = rt + 4
    put(ws, GT, 1, S["etf_global"], F_SECTION, FILL_BLUE, align=A_L)
    ws.merge_cells(start_row=GT, start_column=1, end_row=GT, end_column=9)
    cats_glob = Reference(ws, min_col=1, min_row=r0, max_row=ge)
    barG = BarChart()
    barG.title = S["etf_chart_gl_bar"]
    barG.height, barG.width = 7, 12
    barG.add_data(Reference(ws, min_col=8, min_row=3, max_row=ge), titles_from_data=True)
    barG.set_categories(cats_glob)
    ws.add_chart(barG, "A{}".format(GT + 1))
    pieG = PieChart()
    pieG.title = S["etf_chart_gl_pie"]
    pieG.height, pieG.width = 7, 9
    pieG.add_data(Reference(ws, min_col=7, min_row=r0, max_row=ge), titles_from_data=False)
    pieG.set_categories(cats_glob)
    ws.add_chart(pieG, "F{}".format(GT + 1))

    # --- Interactive per-year detail (dropdown) ---
    SEC = GT + 17
    put(ws, SEC, 1, S["etf_year_section"], F_SECTION, FILL_BLUE, align=A_L)
    ws.merge_cells(start_row=SEC, start_column=1, end_row=SEC, end_column=6)

    SEL = SEC + 1
    put(ws, SEL, 1, S["etf_year_label"], F_LABEL, FILL_LIGHT, align=A_L, border=B_ALL)
    cell_sel = put(ws, SEL, 2, int(years[-1]) if years else None,
                   F_INPUT, FILL_YELLOW, None, A_CV, B_ALL)
    if years:
        dv = DataValidation(type="list", allow_blank=False,
                            formula1='"%s"' % ",".join(str(y) for y in years))
        ws.add_data_validation(dv)
        dv.add(cell_sel)
    ref_year = "$B${}".format(SEL)

    put(ws, SEL + 1, 1, S["etf_by_line"], F_LABEL, align=A_L)
    HDR = SEL + 2
    for j, txt in enumerate(S["etf_ycols"]):
        put(ws, HDR, 1 + j, txt, F_HDR9, FILL_NAVY, align=A_C, border=B_ALL)
    d0 = HDR + 1
    for i in range(nd):
        r = d0 + i
        name = real_names[i] if i < len(real_names) else None
        put(ws, r, 1, name, F_LABEL, FILL_LIGHT, align=A_L, border=B_ALL)
        put(ws, r, 2, '=IF($A{r}="","",SUMIFS({INV}!$G${a}:$G${b},{INV}!$E${a}:$E${b},$A{r},'
            '{INV}!$B${a}:$B${b},{sel}))'.format(INV=INV, a=a, b=b, r=r, sel=ref_year),
            F_CALC, None, FMT_SHARES, border=B_ALL)
        put(ws, r, 3, '=IF($A{r}="","",SUMIFS({INV}!$H${a}:$H${b},{INV}!$E${a}:$E${b},$A{r},'
            '{INV}!$B${a}:$B${b},{sel}))'.format(INV=INV, a=a, b=b, r=r, sel=ref_year),
            F_CALC, None, FMT_EUR, border=B_ALL)
        put(ws, r, 4, '=IF($A{r}="","",SUMIFS({INV}!$J${a}:$J${b},{INV}!$E${a}:$E${b},$A{r},'
            '{INV}!$B${a}:$B${b},{sel}))'.format(INV=INV, a=a, b=b, r=r, sel=ref_year),
            F_CALC, None, FMT_EUR, border=B_ALL)
        put(ws, r, 5, '=IF($A{r}="","",IF(D{r}=0,"",D{r}-C{r}))'.format(r=r),
            F_CALC, None, FMT_EUR_PL, border=B_ALL)
        put(ws, r, 6, '=IF(OR($A{r}="",C{r}=0,D{r}=0),"",(D{r}-C{r})/C{r})'.format(r=r),
            F_CALC, None, FMT_PCT_PL, border=B_ALL)
    dt = d0 + nd
    put(ws, dt, 1, S["total"], F_TOTAL, align=A_L, border=B_ALL)
    for col, f in [(2, FMT_SHARES), (3, FMT_EUR), (4, FMT_EUR), (5, FMT_EUR_PL)]:
        Lc = get_column_letter(col)
        put(ws, dt, col, "=SUM({L}{a}:{L}{b})".format(L=Lc, a=d0, b=dt - 1),
            F_TOTAL, None, f, border=B_ALL)
    put(ws, dt, 6, '=IF(C{dt}=0,"",(D{dt}-C{dt})/C{dt})'.format(dt=dt),
        F_TOTAL, None, FMT_PCT_PL, border=B_ALL)
    colour_pl(ws, "E{a}:F{b}".format(a=d0, b=dt))

    cats = Reference(ws, min_col=1, min_row=d0, max_row=dt - 1)
    bar = BarChart()
    bar.title = S["etf_chart_y_bar"]
    bar.height, bar.width = 7, 12
    bar.add_data(Reference(ws, min_col=5, min_row=HDR, max_row=dt - 1), titles_from_data=True)
    bar.set_categories(cats)
    ws.add_chart(bar, "A{}".format(dt + 2))
    pie = PieChart()
    pie.title = S["etf_chart_y_pie"]
    pie.height, pie.width = 7, 9
    pie.add_data(Reference(ws, min_col=4, min_row=d0, max_row=dt - 1), titles_from_data=False)
    pie.set_categories(cats)
    ws.add_chart(pie, "F{}".format(dt + 2))

    # Monthly detail for the selected year
    MSEC = dt + 18
    put(ws, MSEC, 1, S["etf_month_section"], F_LABEL, align=A_L)
    MHDR = MSEC + 1
    for j, txt in enumerate(S["etf_mcols"]):
        put(ws, MHDR, 1 + j, txt, F_HDR9, FILL_NAVY, align=A_C, border=B_ALL)
    m0 = MHDR + 1
    for i, mn in enumerate(MONTHS):
        r = m0 + i
        put(ws, r, 1, mn, F_LABEL, FILL_LIGHT, align=A_L, border=B_ALL)
        put(ws, r, 2, '=SUMIFS({INV}!$H${a}:$H${b},{INV}!$B${a}:$B${b},{sel},'
            '{INV}!$C${a}:$C${b},"{mn}",{INV}!$D${a}:$D${b},"{buy}")'.format(
                INV=INV, a=a, b=b, sel=ref_year, mn=mn, buy=buy), F_CALC, None, FMT_EUR, border=B_ALL)
        put(ws, r, 3, '=SUMIFS({INV}!$H${a}:$H${b},{INV}!$B${a}:$B${b},{sel},'
            '{INV}!$C${a}:$C${b},"{mn}",{INV}!$D${a}:$D${b},"{sb}")'.format(
                INV=INV, a=a, b=b, sel=ref_year, mn=mn, sb=sb), F_CALC, None, FMT_EUR, border=B_ALL)
        put(ws, r, 4, "=B{r}+C{r}".format(r=r), F_CALC, None, FMT_EUR, border=B_ALL)
        put(ws, r, 5, '=SUMIFS({INV}!$J${a}:$J${b},{INV}!$B${a}:$B${b},{sel},'
            '{INV}!$C${a}:$C${b},"{mn}")'.format(INV=INV, a=a, b=b, sel=ref_year, mn=mn),
            F_CALC, None, FMT_EUR, border=B_ALL)
        put(ws, r, 6, "=E{r}-D{r}".format(r=r), F_CALC, None, FMT_EUR_PL, border=B_ALL)
    mt = m0 + 12
    put(ws, mt, 1, S["total"], F_TOTAL, align=A_L, border=B_ALL)
    for col in (2, 3, 4, 5):
        Lc = get_column_letter(col)
        put(ws, mt, col, "=SUM({L}{a}:{L}{b})".format(L=Lc, a=m0, b=mt - 1),
            F_TOTAL, None, FMT_EUR, border=B_ALL)
    put(ws, mt, 6, "=SUM(F{a}:F{b})".format(a=m0, b=mt - 1), F_TOTAL, None, FMT_EUR_PL, border=B_ALL)
    colour_pl(ws, "F{a}:F{b}".format(a=m0, b=mt))

    barM = BarChart()
    barM.title = S["etf_chart_m_bar"]
    barM.height, barM.width = 7, 16
    barM.add_data(Reference(ws, min_col=6, min_row=MHDR, max_row=mt - 1), titles_from_data=True)
    barM.set_categories(Reference(ws, min_col=1, min_row=m0, max_row=mt - 1))
    ws.add_chart(barM, "A{}".format(mt + 2))
    return rt


# ==========================================================================
#  Dashboard sheet
# ==========================================================================

def build_dashboard(ws, L, entries):
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 3, "B": 15, "C": 9, "D": 9, "E": 9, "F": 11, "G": 9,
                "H": 11, "I": 9, "J": 13, "K": 9, "L": 11, "M": 9, "N": 11, "O": 9})
    tot = L["tot"]
    H = "{INV}!H{r}".format(INV=INV, r=tot)
    J = "{INV}!J{r}".format(INV=INV, r=tot)
    Kf = "{INV}!K{r}".format(INV=INV, r=tot)
    Hbuy = "{INV}!H{r}".format(INV=INV, r=tot + 1)
    Jsav = "{INV}!J{r}".format(INV=INV, r=tot + 1)
    Jsav_val = "{INV}!J{r}".format(INV=INV, r=tot + 2)

    put(ws, 2, 2, S["db_title"], F_TITLE_XL, align=A_L)
    ws.merge_cells("B2:O2")

    # --- KPI cards (7) ---
    cards = [
        (S["db_cards"][0], "=" + Hbuy, FMT_EUR),
        (S["db_cards"][1], "=" + Jsav, FMT_EUR),
        (S["db_cards"][2], "=" + H, FMT_EUR),
        (S["db_cards"][3], "=" + Kf, FMT_EUR),
        (S["db_cards"][4], "=" + J, FMT_EUR),
        (S["db_cards"][5], '=IF(OR({H}=0,{J}=0),"",{J}-{H})'.format(H=H, J=J), FMT_EUR_PL),
        (S["db_cards"][6], '=IF(OR({H}=0,{J}=0),"",({J}-{H})/{H})'.format(H=H, J=J), FMT_PCT_PL),
    ]
    for i, (title, formula, fmt) in enumerate(cards):
        c = 2 + i * 2
        Lc, Lc2 = get_column_letter(c), get_column_letter(c + 1)
        put(ws, 4, c, title, F_HDR9, FILL_NAVY, align=A_C, border=B_ALL)
        ws.merge_cells("{a}4:{b}4".format(a=Lc, b=Lc2))
        put(ws, 5, c, formula, F_KPI, FILL_WHITE, fmt, A_C, B_ALL)
        ws.merge_cells("{a}5:{b}5".format(a=Lc, b=Lc2))
    ws.row_dimensions[4].height = 26
    ws.row_dimensions[5].height = 20

    # --- Performance ---
    put(ws, 7, 2, S["db_perf"], F_SECTION, FILL_BLUE, align=A_L)
    ws.merge_cells("B7:K7")
    for i, txt in enumerate(S["db_perf_cols"]):
        c = 2 + i * 2
        Lc, Lc2 = get_column_letter(c), get_column_letter(c + 1)
        put(ws, 8, c, txt, F_HDR9, FILL_NAVY, align=A_C, border=B_ALL)
        ws.merge_cells("{a}8:{b}8".format(a=Lc, b=Lc2))

    rows_perf = [
        (S["db_perf_tr"], "=" + H,
         '=IF({J}=0,"",{J}-{H})'.format(J=J, H=H),
         '=IF(OR({J}=0,{H}=0),"",({J}-{H})/{H})'.format(J=J, H=H),
         S["db_perf_tr_m"], FILL_GOLD),
        (S["db_perf_sav"], "—",
         '=IF({Jv}=0,"",{Jv})'.format(Jv=Jsav_val),
         '=IF(OR({Jv}=0,{Js}=0),"",({Jv}-{Js})/{Js})'.format(Jv=Jsav_val, Js=Jsav),
         S["db_perf_sav_m"], FILL_LIGHT),
    ]
    for i, (lab, base, ple, plp, measure, flab) in enumerate(rows_perf):
        r = 9 + i
        put(ws, r, 2, lab, F_LABEL if flab is FILL_LIGHT else F_GOLD, flab, align=A_C, border=B_ALL)
        ws.merge_cells("B{r}:C{r}".format(r=r))
        put(ws, r, 4, base, F_CALC, None, FMT_EUR if base != "—" else None, A_C, B_ALL)
        ws.merge_cells("D{r}:E{r}".format(r=r))
        put(ws, r, 6, ple, F_CALC, None, FMT_EUR_PL, A_C, B_ALL)
        ws.merge_cells("F{r}:G{r}".format(r=r))
        put(ws, r, 8, plp, F_CALC, None, FMT_PCT_PL, A_C, B_ALL)
        ws.merge_cells("H{r}:I{r}".format(r=r))
        put(ws, r, 10, measure, F_NOTE, None, None, A_LW, B_ALL)
        ws.merge_cells("J{r}:K{r}".format(r=r))
        ws.row_dimensions[r].height = 24

    put(ws, 12, 2, S["db_perf_note"], F_NOTE, align=A_LW)
    ws.merge_cells("B12:K12")
    colour_pl(ws, "L5", "N5", "F9:G10", "H9:I10")

    # --- Monthly detail: portfolio value over time ---
    put(ws, 15, 2, S["db_month_section"], F_SECTION, FILL_BLUE, align=A_L)
    ws.merge_cells("B15:I15")
    for j, txt in enumerate(S["db_mcols"]):
        put(ws, 16, 2 + j, txt, F_HDR9, FILL_NAVY, align=A_C, border=B_ALL)

    a, b = L["first"], L["last"]
    months = covered_months(entries)
    values = monthly_historic_value(entries, months)
    buy, sb = S["buy"], S["saveback"]
    r0 = 17
    for i, (y, m) in enumerate(months):
        r = r0 + i
        mn = MONTHS[m - 1]
        last = (i == len(months) - 1)
        put(ws, r, 2, "%s %d" % (MONTHS_ABBR[m - 1], y), F_LABEL, FILL_LIGHT, align=A_CV, border=B_ALL)
        put(ws, r, 3, '=SUMIFS({INV}!$H${a}:$H${b},{INV}!$B${a}:$B${b},{y},'
            '{INV}!$C${a}:$C${b},"{mn}",{INV}!$D${a}:$D${b},"{buy}")'.format(
                INV=INV, a=a, b=b, y=y, mn=mn, buy=buy), F_CALC, None, FMT_EUR, border=B_ALL)
        put(ws, r, 4, '=SUMIFS({INV}!$H${a}:$H${b},{INV}!$B${a}:$B${b},{y},'
            '{INV}!$C${a}:$C${b},"{mn}",{INV}!$D${a}:$D${b},"{sb}")'.format(
                INV=INV, a=a, b=b, y=y, mn=mn, sb=sb), F_CALC, None, FMT_EUR, border=B_ALL)
        put(ws, r, 5, "=C{r}+D{r}".format(r=r), F_CALC, None, FMT_EUR, border=B_ALL)
        if i == 0:
            put(ws, r, 6, "=E{r}".format(r=r), F_CALC, None, FMT_EUR, border=B_ALL)
        else:
            put(ws, r, 6, "=F{p}+E{r}".format(p=r - 1, r=r), F_CALC, None, FMT_EUR, border=B_ALL)
        # Mark-to-market value at the month's transaction prices; last point = current
        # value (fetched/typed price), to pin the workbook to its generation date.
        if last:
            put(ws, r, 7, "={INV}!J{tot}".format(INV=INV, tot=L["tot"]),
                F_CALC, None, FMT_EUR, border=B_ALL)
        else:
            put(ws, r, 7, values[i], F_CALC, None, FMT_EUR, border=B_ALL)
        put(ws, r, 8, "=G{r}-F{r}".format(r=r), F_CALC, None, FMT_EUR_PL, border=B_ALL)
        put(ws, r, 9, '=IF(F{r}=0,"",(G{r}-F{r})/F{r})'.format(r=r),
            F_CALC, None, FMT_PCT_PL, border=B_ALL)

    rt = r0 + len(months)
    put(ws, rt, 2, S["total"], F_TOTAL, align=A_CV, border=B_ALL)
    for col in (3, 4):
        Lc = get_column_letter(col)
        put(ws, rt, col, "=SUM({L}{a}:{L}{b})".format(L=Lc, a=r0, b=rt - 1),
            F_TOTAL, None, FMT_EUR, border=B_ALL)
    put(ws, rt, 5, "=C{rt}+D{rt}".format(rt=rt), F_TOTAL, None, FMT_EUR, border=B_ALL)
    put(ws, rt, 6, "=F{p}".format(p=rt - 1), F_TOTAL, None, FMT_EUR, border=B_ALL)   # final cumulative cost
    put(ws, rt, 7, "=G{p}".format(p=rt - 1), F_TOTAL, None, FMT_EUR, border=B_ALL)   # final value
    put(ws, rt, 8, "=G{rt}-F{rt}".format(rt=rt), F_TOTAL, None, FMT_EUR_PL, border=B_ALL)
    put(ws, rt, 9, '=IF(F{rt}=0,"",(G{rt}-F{rt})/F{rt})'.format(rt=rt),
        F_TOTAL, None, FMT_PCT_PL, border=B_ALL)
    colour_pl(ws, "H{a}:H{b}".format(a=r0, b=rt))

    put(ws, rt + 1, 2, S["db_month_note"], F_NOTE, align=A_LW)
    ws.merge_cells("B{r}:I{r}".format(r=rt + 1))

    # --- Charts ---
    if len(months) >= 1:
        cats = Reference(ws, min_col=2, min_row=r0, max_row=rt - 1)
        line = LineChart()
        line.title = S["db_chart_line"]
        line.height, line.width = 8, 16
        line.add_data(Reference(ws, min_col=6, max_col=7, min_row=16, max_row=rt - 1),
                      titles_from_data=True)
        line.set_categories(cats)
        ws.add_chart(line, "B{}".format(rt + 3))
        bar = BarChart()
        bar.title = S["db_chart_bar"]
        bar.height, bar.width = 8, 16
        bar.add_data(Reference(ws, min_col=8, min_row=16, max_row=rt - 1), titles_from_data=True)
        bar.set_categories(cats)
        ws.add_chart(bar, "B{}".format(rt + 20))
        pie = PieChart()
        pie.title = S["db_chart_pie"]
        pie.height, pie.width = 8, 10
        etf_sheet = ws.parent[S["sheet_by_etf"]]
        labels = Reference(etf_sheet, min_col=1, min_row=4, max_row=3 + len(L["names"]))
        vals = Reference(etf_sheet, min_col=7, min_row=4, max_row=3 + len(L["names"]))
        pie.add_data(vals, titles_from_data=False)
        pie.set_categories(labels)
        ws.add_chart(pie, "L{}".format(rt + 3))


# ==========================================================================
#  Tax sheet
# ==========================================================================

def build_tax(ws, tax):
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 3, "B": 16, "C": 16, "D": 16, "E": 16, "F": 18, "G": 16})
    put(ws, 2, 2, S["tax_title"], F_TITLE_XL, align=A_L)
    put(ws, 3, 2, S["tax_sub"], F_SUB, align=A_L)

    for j, txt in enumerate(S["tax_cols"]):
        put(ws, 5, 2 + j, txt, F_HDR, FILL_NAVY, align=A_C, border=B_ALL)

    r0 = 6
    years = list(tax.keys())
    for i, yr in enumerate(years):
        r = r0 + i
        d = tax[yr]
        put(ws, r, 2, int(yr), F_LABEL, FILL_LIGHT, align=A_CV, border=B_ALL)
        put(ws, r, 3, round(d["interest"], 2), F_CALC, None, FMT_EUR_PL, border=B_ALL)
        put(ws, r, 4, round(d["gain"], 2), F_CALC, None, FMT_EUR_PL, border=B_ALL)
        put(ws, r, 5, "=C{r}+D{r}".format(r=r), F_CALC, None, FMT_EUR_PL, border=B_ALL)
        put(ws, r, 6, "=MAX(D{r},0)*0.3+C{r}*0.3".format(r=r), F_CALC, None, FMT_EUR, border=B_ALL)
        put(ws, r, 7, round(d["contrib"], 2), F_CALC, None, FMT_EUR, border=B_ALL)

    rt = r0 + len(years)
    put(ws, rt, 2, S["total"], F_TOTAL, align=A_CV, border=B_ALL)
    for col, f in [(3, FMT_EUR_PL), (4, FMT_EUR_PL), (5, FMT_EUR_PL), (6, FMT_EUR), (7, FMT_EUR)]:
        Lc = get_column_letter(col)
        put(ws, rt, col, "=SUM({L}{a}:{L}{b})".format(L=Lc, a=r0, b=rt - 1) if years else 0,
            F_TOTAL, None, f, border=B_ALL)
    if years:
        colour_pl(ws, "D{a}:D{b}".format(a=r0, b=rt))

    r = rt + 2
    for note in S["tax_notes"]:
        put(ws, r, 2, "• " + note, F_NOTE, align=A_LW)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        r += 1


# ==========================================================================
#  Stock Picking-IPO sheet (single stocks, weighted-average cost)
# ==========================================================================

def build_stock(ws, trades, state, real_isin, by_isin, by_name):
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 24, "B": 14, "C": 13, "D": 12, "E": 10, "F": 13,
                "G": 10, "H": 13, "I": 14, "J": 13})
    put(ws, 1, 1, S["sp_title"], F_TITLE, align=A_L)
    ws.merge_cells("A1:J1")
    put(ws, 2, 1, S["sp_sub"], F_SUB, align=A_L)

    stocks = sorted(state.values(), key=lambda s: s["name"])

    # --- Current prices ---
    put(ws, 4, 1, S["sp_prices"], F_SECTION, FILL_BLUE, align=A_L)
    ws.merge_cells("A4:D4")
    for j, txt in enumerate(S["sp_price_cols"]):
        put(ws, 5, 1 + j, txt, F_HDR, FILL_NAVY, align=A_C, border=B_ALL)
    PRICE_FIRST = 6
    n_slots = max(len(stocks) + 2, 4)
    for i in range(n_slots):
        r = PRICE_FIRST + i
        s = stocks[i] if i < len(stocks) else None
        name = s["name"] if s else None
        isin = s["isin"] if s else None
        price, updated = price_for(isin, name, by_isin, by_name) if name else (None, None)
        put(ws, r, 1, name, F_LABEL, FILL_LIGHT, align=A_L, border=B_ALL)
        put(ws, r, 2, safe_text(isin), F_CALC, FILL_LIGHT, None, A_CV, B_ALL)
        put(ws, r, 3, price, F_INPUT, FILL_YELLOW, FMT_PRICE, border=B_ALL)
        put(ws, r, 4, updated, F_INPUT, FILL_YELLOW, FMT_DATE, border=B_ALL)
    price_range = "$A${a}:$C${b}".format(a=PRICE_FIRST, b=PRICE_FIRST + n_slots - 1)

    # --- Transactions ---
    JHDR = PRICE_FIRST + n_slots + 1
    put(ws, JHDR - 1, 1, S["sp_journal"], F_SECTION, FILL_BLUE, align=A_L)
    ws.merge_cells(start_row=JHDR - 1, start_column=1, end_row=JHDR - 1, end_column=9)
    for j, txt in enumerate(S["sp_jcols"]):
        put(ws, JHDR, 1 + j, txt, F_HDR9, FILL_NAVY, align=A_C, border=B_ALL)
    r = JHDR + 1
    for t in sorted(trades, key=lambda x: x["date"]):
        typ = S["sell"] if t["dir"] == "SELL" else S["buy"]
        put(ws, r, 1, t["date"], F_CALC, None, FMT_DATE, border=B_ALL)
        put(ws, r, 2, typ, F_CALC, None, None, A_CV, B_ALL)
        put(ws, r, 3, t["name"], F_CALC, None, None, A_L, B_ALL)
        put(ws, r, 4, safe_text(t["isin"]), F_CALC, None, None, A_L, B_ALL)
        put(ws, r, 5, t["qty"], F_CALC, None, FMT_SHARES, border=B_ALL)
        put(ws, r, 6, t["price"], F_CALC, None, FMT_PRICE, border=B_ALL)
        put(ws, r, 7, t["fee"], F_CALC, None, FMT_EUR, border=B_ALL)
        put(ws, r, 8, (t["amount"] if t["dir"] == "SELL" else -t["amount"]),
            F_CALC, None, FMT_EUR, border=B_ALL)
        put(ws, r, 9, typ, F_CALC, None, None, A_L, B_ALL)
        r += 1
    if not trades:
        put(ws, r, 1, S["sp_empty"], F_NOTE, align=A_L)
        r += 1

    # --- Summary per stock ---
    SHDR = r + 2
    put(ws, SHDR - 1, 1, S["sp_summary"], F_SECTION, FILL_BLUE, align=A_L)
    ws.merge_cells(start_row=SHDR - 1, start_column=1, end_row=SHDR - 1, end_column=10)
    for j, txt in enumerate(S["sp_scols"]):
        put(ws, SHDR, 1 + j, txt, F_HDR9, FILL_NAVY, align=A_C, border=B_ALL)
    r = SHDR + 1
    for s in stocks:
        wac = s["cost"] / s["qty"] if s["qty"] else 0
        put(ws, r, 1, s["name"], F_LABEL, FILL_LIGHT, align=A_L, border=B_ALL)
        put(ws, r, 2, round(s["qty"], 6), F_CALC, None, FMT_SHARES, border=B_ALL)
        put(ws, r, 3, round(wac, 4), F_CALC, None, FMT_PRICE, border=B_ALL)
        put(ws, r, 4, round(s["cost"], 2), F_CALC, None, FMT_EUR, border=B_ALL)
        put(ws, r, 5, '=IFERROR(VLOOKUP($A{r},{rg},3,0),"")'.format(r=r, rg=price_range),
            F_CALC, None, FMT_PRICE, border=B_ALL)
        put(ws, r, 6, '=IF(OR(B{r}="",E{r}=""),"",B{r}*E{r})'.format(r=r),
            F_CALC, None, FMT_EUR, border=B_ALL)
        put(ws, r, 7, '=IF(F{r}="","",F{r}-D{r})'.format(r=r), F_CALC, None, FMT_EUR_PL, border=B_ALL)
        put(ws, r, 8, '=IF(OR(D{r}=0,F{r}=""),"",(F{r}-D{r})/D{r})'.format(r=r),
            F_CALC, None, FMT_PCT_PL, border=B_ALL)
        put(ws, r, 9, round(real_isin.get(s["isin"], 0.0), 2), F_CALC, None, FMT_EUR_PL, border=B_ALL)
        r += 1
    if stocks:
        colour_pl(ws, "G{a}:I{b}".format(a=SHDR + 1, b=r - 1))
    else:
        put(ws, r, 1, "—", F_NOTE, align=A_L)

    ws.freeze_panes = "A6"


# ==========================================================================
#  Price preservation and auto-fetch
# ==========================================================================

def read_prices(path):
    """Re-read prices already typed into an existing --fo, by ISIN and by name.

    Locates the table via the "Current price" / "Cours actuel" header, so it is
    robust to column order and compatible with either language.
    """
    by_isin, by_name = {}, {}
    if not path or not os.path.exists(path):
        return by_isin, by_name
    try:
        wb = openpyxl.load_workbook(path)
    except Exception:
        return by_isin, by_name
    price_tokens = ("Current price", "Cours actuel")
    name_tokens = ("ETF", "Company", "Société", "action", "stock")
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            hdr = {c.column: c.value for c in row if isinstance(c.value, str)}
            price_col = next((k for k, v in hdr.items()
                              if any(t in v for t in price_tokens)), None)
            if not price_col:
                continue
            name_col = next((k for k, v in hdr.items()
                             if any(t in v for t in name_tokens)), None)
            isin_col = next((k for k, v in hdr.items() if "ISIN" in v), None)
            if not name_col:
                break
            r = row[0].row + 1
            while True:
                name = ws.cell(r, name_col).value
                if not isinstance(name, str) or not name.strip():
                    break
                price = ws.cell(r, price_col).value
                updated = ws.cell(r, price_col + 1).value
                isin = ws.cell(r, isin_col).value if isin_col else None
                if isinstance(price, (int, float)):
                    by_name[norm_name(name)] = (price, updated)
                    if isin:
                        by_isin[str(isin)] = (price, updated)
                r += 1
            break
    return by_isin, by_name


def price_for(isin, name, by_isin, by_name):
    """Preserved price: ISIN first, fall back to the normalised name."""
    if isin and isin in by_isin:
        return by_isin[isin]
    return by_name.get(norm_name(name), (None, None))


# Public sources queried by ISIN, in order:
#   1. Deutsche Boerse / Xetra (api.boerse-frankfurt.de): European reference
#      venue, EUR quotes, closest to Trade Republic;
#   2. Yahoo Finance: international fallback if the ISIN is not on Xetra.
# If no source answers, the price is left empty (to be filled by hand).

def _http_json(url):
    req = urllib.request.Request(
        url, headers={"User-Agent": "Mozilla/5.0", "Accept": "application/json"})
    with urllib.request.urlopen(req, timeout=12) as r:
        return json.loads(r.read().decode())


def _source_deutsche_boerse(isin):
    """EUR price via Deutsche Boerse / Xetra. None if unavailable."""
    for mic in ("XETR", "XFRA"):
        try:
            d = _http_json("https://api.boerse-frankfurt.de/v1/data/"
                           "quote_box/single?isin=%s&mic=%s" % (isin, mic))
        except Exception:
            continue
        price = d.get("lastPrice")
        currency = d.get("currency")
        if isinstance(currency, dict):
            currency = currency.get("originalValue")
        if price and currency in (None, "EUR"):
            return round(float(price), 2)
    return None


_EXCH_EU = {"GER", "FRA", "STU", "MUN", "HAM", "DUS", "BER", "MIL", "PAR",
            "AMS", "EBS", "MCE", "VIE", "LSE"}


def _yahoo_meta(symbol):
    d = _http_json("https://query1.finance.yahoo.com/v8/finance/chart/" + symbol)
    return d["chart"]["result"][0]["meta"]


def _source_yahoo(isin):
    """EUR price via Yahoo Finance (fallback). None if unavailable."""
    try:
        d = _http_json("https://query1.finance.yahoo.com/v1/finance/search?q=" + isin)
        quotes = d.get("quotes", [])
        if not quotes:
            return None
        eu = [q for q in quotes if q.get("exchange") in _EXCH_EU]
        meta = _yahoo_meta((eu or quotes)[0].get("symbol"))
        price, currency = meta.get("regularMarketPrice"), meta.get("currency")
        if price is None:
            return None
        rate = 1.0 if currency == "EUR" else _yahoo_meta(currency + "EUR=X").get("regularMarketPrice")
        return round(price * rate, 2) if rate else None
    except Exception:
        return None


_SOURCES = (_source_deutsche_boerse, _source_yahoo)


def fetch_price(isin):
    """Current EUR price for an ISIN, trying each source. None on failure."""
    for source in _SOURCES:
        try:
            price = source(isin)
        except Exception:
            price = None
        if price is not None:
            return price
    return None


def fetch_all_prices(isins, now):
    """Return {isin: (eur_price, date)} for resolved ISINs; failures are skipped."""
    result, missing = {}, []
    for isin in sorted(i for i in isins if i):
        price = fetch_price(isin)
        if price is not None:
            result[isin] = (price, now)
            print(S["cli_price_line"] % (isin, price))
        else:
            missing.append(isin)
    if missing:
        print(S["cli_missing"] % ", ".join(missing))
    return result


# ==========================================================================
#  Assembly / CLI
# ==========================================================================

def generate(args):
    """Read the CSV in args.fi and write the workbook to args.fo."""
    select_language(args.lang)

    rows = validate_export(args.fi)                # untrusted input -> validated
    _ISIN_NAMES.update(index_names(rows))          # isin -> name (from the CSV)
    name_isin = {v: k for k, v in _ISIN_NAMES.items()}
    entries = build_etf_journal(rows)
    etf_trades = [to_trade(r) for r in rows
                  if is_trade(r) and r["asset_class"] == "FUND" and r["symbol"]]
    single_trades = stock_trades(rows)
    _, real_etf_year, _ = positions_wac(etf_trades)
    stock_state, real_stock_year, real_stock_isin = positions_wac(single_trades)
    real_year = defaultdict(float)
    for y, g in dict(real_etf_year).items():
        real_year[y] += g
    for y, g in real_stock_year.items():
        real_year[y] += g
    tax = tax_summary(rows, entries, real_year)
    years = sorted({str(e["date"].year) for e in entries} | set(tax.keys()))
    by_isin, by_name = read_prices(args.fo)

    # Auto prices (optional): take precedence over preserved values.
    if args.auto_prices:
        isins = {name_isin.get(e["name"]) for e in entries} | set(stock_state)
        print(S["cli_fetch"])
        by_isin.update(fetch_all_prices(isins, datetime.now()))

    # Safety backup before overwriting.
    if os.path.exists(args.fo):
        shutil.copy2(args.fo, args.fo + ".bak")

    wb = Workbook()
    wb.remove(wb.active)
    ws_rm = wb.create_sheet(S["sheet_readme"])
    ws_db = wb.create_sheet(S["sheet_dashboard"])
    ws_inv = wb.create_sheet(S["sheet_investments"])
    ws_yr = wb.create_sheet(S["sheet_yearly"])
    ws_etf = wb.create_sheet(S["sheet_by_etf"])
    ws_tax = wb.create_sheet(S["sheet_tax"])
    ws_sp = wb.create_sheet(S["sheet_stock"])

    build_readme(ws_rm)
    L = build_investments(ws_inv, entries, name_isin, by_isin, by_name)
    build_yearly(ws_yr, L, years)
    build_by_etf(ws_etf, L, years)
    build_tax(ws_tax, tax)
    build_stock(ws_sp, single_trades, stock_state, real_stock_isin, by_isin, by_name)
    build_dashboard(ws_db, L, entries)             # after By-ETF (pie chart source)

    wb.active = 1                                  # open on the Dashboard
    wb.save(args.fo)

    n_b = sum(1 for e in entries if e["kind"] == "buy")
    n_s = sum(1 for e in entries if e["kind"] == "saveback")
    n_v = sum(1 for e in entries if e["kind"] == "sell")
    print(S["cli_done"] % args.fo)
    print(S["cli_journal"] % (len(entries), n_b, n_s, n_v))
    print(S["cli_etf"] % ", ".join(L["names"]))
    if single_trades:
        print(S["cli_stock"] % (len(single_trades), len(stock_state)))
    for yr, d in tax.items():
        print(S["cli_tax"] % (yr, d["interest"], d["gain"], d["contrib"]))
    if os.path.exists(args.fo + ".bak"):
        print(S["cli_backup"] % os.path.basename(args.fo))


def main():
    p = argparse.ArgumentParser(description=TXT["en"]["cli_desc"])
    p.add_argument("--fi", default="transactions.csv", help="Trade Republic CSV export")
    p.add_argument("--fo", default="TradeRepublicBoard.xlsx", help="Output workbook")
    lang = p.add_mutually_exclusive_group()
    lang.add_argument("--en", action="store_const", dest="lang", const="en",
                      help="Generate the workbook in English (default)")
    lang.add_argument("--fr", action="store_const", dest="lang", const="fr",
                      help="Generate the workbook in French")
    p.add_argument("--auto-prices", action="store_true",
                   help="Fetch current prices by ISIN (Deutsche Börse / Yahoo, needs internet)")
    p.add_argument("--watch", metavar="DIR",
                   help="One-shot watch: if DIR contains the export file, process it "
                        "and delete it on success (meant for cron / Task Scheduler)")
    p.add_argument("--watch-name", default="trade-republic-export.csv",
                   help="Expected export filename inside --watch DIR")
    p.set_defaults(lang="en")
    args = p.parse_args()

    if args.watch:
        src = os.path.join(args.watch, args.watch_name)
        if not os.path.exists(src):
            return                                 # nothing to do, exit quietly
        args.fi = src
        generate(args)                             # raises on failure -> file kept
        os.remove(src)                             # delete only after a successful run
    else:
        generate(args)


if __name__ == "__main__":
    main()
